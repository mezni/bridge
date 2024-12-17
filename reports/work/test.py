import warnings
import pandas as pd
from openpyxl import load_workbook

def list_sheets(wb):
    return wb.sheetnames

def extract_excel_formulas(workbook, sheet_name):
    sheet = workbook[sheet_name] 

    # List to store extracted data
    extracted_data = []

    # Iterate through cells and extract data
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell_data = {
                    "coordinate": cell.coordinate,
                    "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None,
                    "value": cell.value,
                }
                extracted_data.append(cell_data)

    return extracted_data



warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
file_path = 'SuiviDepenses.xlsx'  
workbook  = load_workbook(file_path)
sheet_names=list_sheets(workbook)
extracted_data=extract_excel_formulas(workbook, "Sélection")
#rint (extracted_data)
#for i in extracted_data:
#    if i["formula"]: 
#        print (i["coordinate"],i["formula"]) 


#sheet = workbook["Sélection"] 
#data = [[cell.value for cell in row] for row in sheet.iter_rows()]
#print (workbook.defined_names) 

