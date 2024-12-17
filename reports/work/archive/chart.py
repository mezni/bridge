import pandas as pd
from openpyxl import load_workbook

def list_sheet_names(file_path):
    wb = load_workbook(file_path, read_only=True)
    return wb.sheetnames

# Example usage
file_path = 'sales_02.xlsx'
sheet_names = list_sheet_names(file_path)
print('## Sheets')
for sheet_name in sheet_names:
    print(sheet_name)


###########################################

from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('sales_02.xlsx')

# Iterate through all sheets in the workbook
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Check if there are any charts in the sheet
    if sheet._charts:
        print(f"Charts in sheet: {sheet_name}")
        
        # Iterate over each chart in the sheet
        for chart in sheet._charts:
            chart_type = type(chart).__name__  # Get the type of the chart (e.g., BarChart, LineChart)
            print(f"  Chart Type: {chart_type}")
            
            # Accessing chart title
            if chart.title:
                title_text = ""
                if chart.title.tx.rich:
                    for paragraph in chart.title.tx.rich.p:
                        for run in paragraph.r:
                            title_text += run.t  # Collecting all text runs
                print(f"  Chart Title: {title_text if title_text else 'No title text'}")
            
            # Accessing data (data series)
            if chart.series:
                for idx, series in enumerate(chart.series):
                    print(f"    Series {idx + 1}:")
                    print(series)
#                    print(f"      Data Reference: {series.values}")
                    if hasattr(series, 'categories'):
                        print(f"      Categories Reference: {series.categories}")
                    else:
                        print(f"      No categories reference")
            
            # Accessing x-axis and y-axis titles (if available)
            if chart.x_axis.title:
                print(f"  X-Axis Title: {chart.x_axis.title}")
            else:
                print(f"  No X-Axis Title")
            
            if chart.y_axis.title:
                print(f"  Y-Axis Title: {chart.y_axis.title}")
            else:
                print(f"  No Y-Axis Title")
            
            # Additional properties
            print(f"  Chart Style: {chart.style}")
#            print(f"  Chart Legend: {'Visible' if chart.has_legend else 'Hidden'}")
            print(f"  Plot Area: {chart.plot_area}")
            print(f"  Chart Width: {chart.width}, Chart Height: {chart.height}")
            
    else:
        print(f"No charts found in sheet: {sheet_name}")
