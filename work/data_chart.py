import pandas as pd
from openpyxl import load_workbook

def check_sheets_for_data_and_pivot(file_path):
    # Load the Excel file
    wb = load_workbook(file_path, keep_links=False)
    
    # List all sheet names
    sheet_names = wb.sheetnames
    print(f"Sheets in the file: {sheet_names}")
    
    # Iterate through each sheet
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        print(f"Checking sheet: {sheet_name}")
        
        # Check if the sheet has any data (non-empty rows and columns)
        has_data = False
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                has_data = True
                break
        
        if has_data:
            print(f"  - Data found in sheet {sheet_name}")
        else:
            print(f"  - No data in sheet {sheet_name}")
        
        # Check for pivot tables
        has_pivot = False
        for pivot in wb._pivots:
            if pivot and pivot.sheet == sheet:
                has_pivot = True
                break
        
        if has_pivot:
            print(f"  - Pivot table found in sheet {sheet_name}")
        else:
            print(f"  - No pivot table in sheet {sheet_name}")
        
        # Check for charts (graphs)
        charts = sheet._charts
        if charts:
            print(f"  - Graphs found in sheet {sheet_name}")
        else:
            print(f"  - No graphs in sheet {sheet_name}")

# Example usage
file_path = 'sales.xlsx'  # Change this to the path of your Excel file
check_sheets_for_data_and_pivot(file_path)





import pandas as pd
from openpyxl import load_workbook

def check_sheets_for_data_and_charts(file_path):
    # Load the Excel file
    wb = load_workbook(file_path, keep_links=False)
    
    # List all sheet names
    sheet_names = wb.sheetnames
    print(f"Sheets in the file: {sheet_names}")
    
    # Iterate through each sheet
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        print(f"Checking sheet: {sheet_name}")
        
        # Check if the sheet has any data (non-empty rows and columns)
        has_data = False
        data = []  # List to hold rows of data
        for row in sheet.iter_rows(values_only=True):
            # If any cell in the row is not None, consider it as data
            if any(cell is not None for cell in row):
                has_data = True
                data.append(row)  # Add the row to data

        if has_data:
            print(f"  - Data found in sheet {sheet_name}")
            # Convert the data to a pandas DataFrame for better formatting
            df = pd.DataFrame(data)
            print(df.to_string(index=False))  # Print the data without index
        else:
            print(f"  - No data in sheet {sheet_name}")
        
        # Check for charts (graphs)
        charts = sheet._charts
        if charts:
            print(f"  - Graphs found in sheet {sheet_name}")
        else:
            print(f"  - No graphs in sheet {sheet_name}")

# Example usage
file_path = 'sales.xlsx'  # Change this to the path of your Excel file
check_sheets_for_data_and_charts(file_path)



import pandas as pd
from openpyxl import load_workbook

def get_chart_title(chart):
    if chart.title:
        if isinstance(chart.title.text, str):
            return chart.title.text  # Directly access the text of the title
    return None

def check_sheets_for_data_and_charts(file_path):
    # Load the Excel file
    wb = load_workbook(file_path, keep_links=False)
    
    # List all sheet names
    sheet_names = wb.sheetnames
    print(f"Sheets in the file: {sheet_names}")
    
    # Iterate through each sheet
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        print(f"Checking sheet: {sheet_name}")
        
        # Check if the sheet has any data (non-empty rows and columns)
        has_data = False
        data = []  # List to hold rows of data
        for row in sheet.iter_rows(values_only=True):
            # If any cell in the row is not None, consider it as data
            if any(cell is not None for cell in row):
                has_data = True
                data.append(row)  # Add the row to data

        if has_data:
            print(f"  - Data found in sheet {sheet_name}")
            # Convert the data to a pandas DataFrame for better formatting
            df = pd.DataFrame(data)
            print(df.to_string(index=False))  # Print the data without index
        else:
            print(f"  - No data in sheet {sheet_name}")
        
        # Check for charts (graphs)
        charts = sheet._charts
        if charts:
            print(f"  - Graphs found in sheet {sheet_name}")
            for chart in charts:
                print(f"    - Chart Type: {chart.type}")  # Print chart type (Bar, Line, etc.)
                
                # Get and print chart title
                title = get_chart_title(chart)
                if title:
                    print(f"    - Chart Title: {title}")  # Print the chart title
                else:
                    print(f"    - No title for this chart")
                
                print(f"    - Number of Data Series: {len(chart.series)}")  # Number of data series in the chart
        else:
            print(f"  - No graphs in sheet {sheet_name}")

# Example usage
file_path = 'sales.xlsx'  # Change this to the path of your Excel file
check_sheets_for_data_and_charts(file_path)
