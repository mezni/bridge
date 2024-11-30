import warnings
import pandas as pd
from openpyxl import load_workbook

def list_sheets(wb):
    return wb.sheetnames


def check_sheet_for_data(workbook, sheet_name):
    # Get the sheet
    sheet = workbook[sheet_name]
    print(f"Checking sheet: {sheet_name}")
    
    # Check for data (non-empty rows and columns)
    has_data = False
    data = []  # List to hold rows of data
    
    # Iterate through rows and check for non-empty cells
    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None for cell in row):  # Check if there's any non-empty cell
            has_data = True
            data.append(row)  # Add the row to the data list

    # Print results
    if has_data:
        print(f"  - Data found in sheet {sheet_name}")
        # Convert the data to a pandas DataFrame for better formatting
        df = pd.DataFrame(data)
        print(df.to_string(index=False))  # Print the data without the index
    else:
        print(f"  - No data in sheet {sheet_name}")

def detect_charts_in_sheet(workbook, sheet_name):

    
    sheet = workbook[sheet_name]
    print(f"Checking charts in sheet: {sheet_name}")
    
    # Get all charts in the sheet
    charts = sheet._charts
    
    if charts:
        print(f"  - Found {len(charts)} chart(s) in sheet '{sheet_name}':")
        for idx, chart in enumerate(charts, start=1):
            print(f"    Chart {idx}:")
            print(f"      Type: {chart.__class__.__name__}")
            print(f"      Title: {chart.title if chart.has_title else 'No title'}")
            print(f"      Location: {chart.anchor}")
    else:
        print(f"  - No charts found in sheet '{sheet_name}'")

from openpyxl import load_workbook

def detect_segments_in_sheet(workbook, sheet_name):  
    sheet = workbook[sheet_name]
    # Detect merged cells
    merged_cells = sheet.merged_cells.ranges
    if merged_cells:
        print(f"  - Found {len(merged_cells)} merged cell(s):")
        for merged in merged_cells:
            print(f"    Merged range: {merged}")
    else:
        print(f"  - No merged cells found in sheet '{sheet_name}'")
    
    # Detect hidden rows (rows with 'hidden' attribute set to True)
    hidden_rows = []
    for row_idx in range(1, sheet.max_row + 1):
        if sheet.row_dimensions[row_idx].hidden:
            hidden_rows.append(row_idx)
    
    if hidden_rows:
        print(f"  - Found {len(hidden_rows)} hidden row(s) (potential grouping):")
        for row in hidden_rows:
            print(f"    Hidden row: {row}")
    else:
        print(f"  - No hidden rows found in sheet '{sheet_name}'")
    
    # Detect hidden columns (columns with 'hidden' attribute set to True)
    hidden_columns = []
    for col_idx in range(1, sheet.max_column + 1):
        col_letter = chr(64 + col_idx)  # Convert column index to letter (A, B, C, ...)
        if sheet.column_dimensions[col_letter].hidden:
            hidden_columns.append(col_letter)
    
    if hidden_columns:
        print(f"  - Found {len(hidden_columns)} hidden column(s) (potential grouping):")
        for col in hidden_columns:
            print(f"    Hidden column: {col}")
    else:
        print(f"  - No hidden columns found in sheet '{sheet_name}'")



warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
file_path = 'SuiviDepenses.xlsx'  
workbook  = load_workbook(file_path)
sheet_names=list_sheets(workbook)
print(f"Sheets: {sheet_names}")
check_sheet_for_data(workbook, "Sélection")
detect_charts_in_sheet(workbook, "Sélection")
detect_segments_in_sheet(workbook, "Sélection")



def print_worksheet_metadata(sheet):
    metadata = {}

    # Basic information
    metadata['sheet_name'] = sheet.title
    metadata['num_rows'] = sheet.max_row
    metadata['num_columns'] = sheet.max_column

    # Column headers (first row)
    headers = [cell.value for cell in sheet[1]]  # Assumes first row contains headers
    metadata['headers'] = headers

    # Merged cells
    merged_cells = [str(merge) for merge in sheet.merged_cells.ranges]
    metadata['merged_cells'] = merged_cells

    # Data types of the first few rows for each column
    column_data_types = {}
    for col_idx, header in enumerate(headers, start=1):
        column_data_types[header] = set()  # Use set to store unique data types
        
        for row in range(2, sheet.max_row + 1):  # Start from row 2 to skip the header
            cell_value = sheet.cell(row=row, column=col_idx).value
            if cell_value is not None:
                column_data_types[header].add(type(cell_value).__name__)  # Store the type name
    
    metadata['column_data_types'] = column_data_types

    # Formulas (if any)
    formulas = {}
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            if cell.data_type == 'f':  # 'f' indicates a formula cell
                formulas[cell.coordinate] = cell.value
    
    metadata['formulas'] = formulas

    # Print the metadata
    for key, value in metadata.items():
        print(f"{key}: {value}")

sheet = workbook["Sélection"]
print_worksheet_metadata(sheet)