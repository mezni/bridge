import pandas as pd
from openpyxl import load_workbook

def list_sheet_names(file_path):
    wb = load_workbook(file_path, read_only=True)
    return wb.sheetnames

# Example usage
file_path = 'sales_02.xlsx'
sheet_names = list_sheet_names(file_path)
print('## Sheets')
for sheet_name in sheet_names:
    print(sheet_name)


###########################################

from openpyxl import load_workbook

# Load the workbook
workbook = load_workbook('sales_02.xlsx')


# Iterate through each sheet
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    
    # Check if there are charts in the sheet
    if sheet._charts:
        print(f"Charts in sheet: {sheet_name}")
        for chart in sheet._charts:
            print(f"Chart type: {type(chart).__name__}")
            print(f"Chart title: {chart.ser}")
    else:
        print(f"No charts in sheet: {sheet_name}")

