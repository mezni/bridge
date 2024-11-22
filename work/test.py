import pandas as pd
from openpyxl import load_workbook

def list_sheet_names(file_path):
    wb = load_workbook(file_path, read_only=True)
    return wb.sheetnames

# Example usage
file_path = 'sales_02.xlsx'
sheet_names = list_sheet_names(file_path)
print('## Sheets')
for sheet_name in sheet_names:
    print(sheet_name)

df = pd.read_excel(file_path, sheet_name='sales')
print( df.head(5))
print( df.tail(5))
print("Number of rows:", df.shape[0])  # Number of rows
print("Number of columns:", df.shape[1])  # Number of columns
df = df.dropna()
print("Number of rows:", df.shape[0])  # Number of rows
print("Number of columns:", df.shape[1])  # Number of columns
print( df.head(5))  # Number of columns
json_data = df.to_json(orient='records', lines=True)
print (json_data)

